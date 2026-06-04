import json
import os
from datetime import timedelta
from io import BytesIO

from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone

from .models import Match, Equipe, Joueur, Participation, Evenement, ScorePeriode, Sponsor, SponsorGlobal


# ==========================================
# 1. ACCUEIL & GESTION ÉQUIPES
# ==========================================

def accueil(request):
    equipes_existantes = Equipe.objects.all()
    sponsors_globaux = SponsorGlobal.objects.all()

    # Dictionnaire {id_equipe: url_logo} pour le JS (preview à la sélection)
    equipes_logos = {}
    for eq in equipes_existantes:
        if eq.logo:
            try:
                equipes_logos[str(eq.id)] = request.build_absolute_uri(eq.logo.url)
            except Exception:
                pass

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'creer_equipe':
            Equipe.objects.create(nom=request.POST.get('nouveau_nom_equipe', '').strip())
            return redirect('accueil')

        elif action == 'supprimer_equipe':
            Equipe.objects.filter(id=request.POST.get('equipe_id')).delete()
            return redirect('accueil')

        elif action == 'supprimer_sponsor_global':
            sg = SponsorGlobal.objects.filter(id=request.POST.get('sponsor_global_id')).first()
            if sg:
                sg.image.delete(save=False)
                sg.delete()
            return redirect('accueil')

        elif action == 'lancer_match':
            # ── Équipes ──────────────────────────────────────────────────────
            id_dom = request.POST.get('select_domicile')
            nom_dom = request.POST.get('input_domicile', '').strip()
            eq_dom_obj = None
            if id_dom:
                eq_dom_obj = Equipe.objects.filter(id=id_dom).first()
                if eq_dom_obj:
                    nom_dom = eq_dom_obj.nom

            id_ext = request.POST.get('select_exterieur')
            nom_ext = request.POST.get('input_exterieur', '').strip()
            eq_ext_obj = None
            if id_ext:
                eq_ext_obj = Equipe.objects.filter(id=id_ext).first()
                if eq_ext_obj:
                    nom_ext = eq_ext_obj.nom

            try:
                duree_periode = int(request.POST.get('duree_periode', 8))
            except (ValueError, TypeError):
                duree_periode = 8

            heure_raw = request.POST.get('heure_debut', '').strip()
            heure_val = None
            if heure_raw:
                try:
                    from datetime import time as dtime
                    h, m = heure_raw.split(':')
                    heure_val = dtime(int(h), int(m))
                except (ValueError, TypeError):
                    pass

            def _post_int(key, default):
                try:
                    return int(request.POST.get(key, default))
                except (ValueError, TypeError):
                    return default

            match = Match.objects.create(
                nom_equipe_domicile=nom_dom,
                equipe_domicile_origine=eq_dom_obj,
                couleur_bonnet_dom=request.POST.get('couleur_bonnet_dom', 'BLANC'),
                nom_equipe_exterieur=nom_ext,
                equipe_exterieur_origine=eq_ext_obj,
                couleur_bonnet_ext=request.POST.get('couleur_bonnet_ext', 'BLEU'),
                lieu=request.POST.get('lieu', 'Piscine'),
                competition=request.POST.get('competition', ''),
                heure_debut=heure_val,
                # Règles
                duree_periode=duree_periode,
                nb_periodes=_post_int('nb_periodes', 4),
                temps_possession=_post_int('temps_possession', 30),
                duree_exclusion=_post_int('duree_exclusion', 20),
                nb_temps_morts=_post_int('nb_temps_morts', 2),
                max_fautes_perso=_post_int('max_fautes_perso', 3),
                nb_remplacants=_post_int('nb_remplacants', 6),
                # Officiels
                arbitre1_nom=request.POST.get('arbitre1_nom', ''),
                arbitre1_iuf=request.POST.get('arbitre1_iuf', ''),
                arbitre2_nom=request.POST.get('arbitre2_nom', ''),
                arbitre2_iuf=request.POST.get('arbitre2_iuf', ''),
                secretaire_nom=request.POST.get('secretaire_nom', ''),
                secretaire_iuf=request.POST.get('secretaire_iuf', ''),
                chrono_nom=request.POST.get('chrono_nom', ''),
                chrono_iuf=request.POST.get('chrono_iuf', ''),
                juge_but1_nom=request.POST.get('juge_but1_nom', ''),
                juge_but1_iuf=request.POST.get('juge_but1_iuf', ''),
                juge_but2_nom=request.POST.get('juge_but2_nom', ''),
                juge_but2_iuf=request.POST.get('juge_but2_iuf', ''),
                delegue_ffn_nom=request.POST.get('delegue_ffn_nom', ''),
                delegue_ffn_iuf=request.POST.get('delegue_ffn_iuf', ''),
                delegue_dom_nom=request.POST.get('delegue_dom_nom', ''),
                delegue_ext_nom=request.POST.get('delegue_ext_nom', ''),
                # Chrono initial
                temps_restant=duree_periode * 60,
                shot_restant=_post_int('temps_possession', 28),  # synchro avec temps_possession
                # Personnalisation scoreboard
                scoreboard_score_scale=_post_int('scoreboard_score_scale', 100),
                scoreboard_sponsor_height=_post_int('scoreboard_sponsor_height', 10),
                scoreboard_players_visible=(
                    request.POST.get('scoreboard_players_visible') == 'true'),
            )

            # ── Logos équipes ─────────────────────────────────────────────────
            # DOM : si un fichier est uploadé, on le sauvegarde sur l'équipe ET le match
            logo_dom_file = request.FILES.get('logo_dom')
            if logo_dom_file:
                if eq_dom_obj:
                    eq_dom_obj.logo = logo_dom_file
                    eq_dom_obj.save(update_fields=['logo'])
                    match.logo_dom = eq_dom_obj.logo.name
                    match.save(update_fields=['logo_dom'])
                else:
                    match.logo_dom = logo_dom_file
                    match.save(update_fields=['logo_dom'])
            elif eq_dom_obj and eq_dom_obj.logo:
                # Pas de nouveau fichier → réutiliser le logo sauvegardé de l'équipe
                match.logo_dom = eq_dom_obj.logo.name
                match.save(update_fields=['logo_dom'])

            # EXT
            logo_ext_file = request.FILES.get('logo_ext')
            if logo_ext_file:
                if eq_ext_obj:
                    eq_ext_obj.logo = logo_ext_file
                    eq_ext_obj.save(update_fields=['logo'])
                    match.logo_ext = eq_ext_obj.logo.name
                    match.save(update_fields=['logo_ext'])
                else:
                    match.logo_ext = logo_ext_file
                    match.save(update_fields=['logo_ext'])
            elif eq_ext_obj and eq_ext_obj.logo:
                match.logo_ext = eq_ext_obj.logo.name
                match.save(update_fields=['logo_ext'])

            # ── Nouveaux sponsors uploadés → sauvegardés globalement ──────────
            next_ordre = SponsorGlobal.objects.count()
            sponsor_match_ordre = 0
            for sponsor_file in request.FILES.getlist('sponsors'):
                sg = SponsorGlobal.objects.create(
                    image=sponsor_file, ordre=next_ordre)
                next_ordre += 1
                # Référencer le même fichier dans le match (sans re-uploader)
                s = Sponsor(match=match, ordre=sponsor_match_ordre)
                s.image = sg.image.name
                s.save()
                sponsor_match_ordre += 1

            # ── Sponsors globaux sélectionnés ─────────────────────────────────
            for sg_id in request.POST.getlist('sponsors_globaux'):
                try:
                    sg = SponsorGlobal.objects.get(id=int(sg_id))
                    s = Sponsor(match=match, ordre=sponsor_match_ordre)
                    s.image = sg.image.name
                    s.save()
                    sponsor_match_ordre += 1
                except (SponsorGlobal.DoesNotExist, ValueError):
                    pass

            return redirect('compo_equipe', match_id=match.id)

    return render(request, 'gestion/accueil.html', {
        'equipes': equipes_existantes,
        'sponsors_globaux': sponsors_globaux,
        'equipes_logos_json': json.dumps(equipes_logos),
    })


# ==========================================
# 2. COMPOSITION D'ÉQUIPE
# ==========================================

def _build_slots(equipe_origine):
    """
    Retourne une liste de 15 dicts (index 0 = bonnet 1, ..., 14 = bonnet 15).
    Compatible avec les templates Django natifs : pas de filtre personnalisé.
    """
    raw = {}
    if equipe_origine:
        for j in equipe_origine.effectif.all():
            if 0 < j.numero_habituel <= 15:
                raw[j.numero_habituel] = j
    slots = []
    for i in range(1, 16):
        j = raw.get(i)
        slots.append({
            'bonnet':          i,
            'nom':             j.nom             if j else '',
            'prenom':          j.prenom          if j else '',
            'numero_licence':  getattr(j, 'numero_licence', '')  if j else '',
            'annee_naissance': getattr(j, 'annee_naissance', '') if j else '',
        })
    return slots


def compo_equipe(request, match_id):
    match = get_object_or_404(Match, id=match_id)

    if request.method == 'POST':
        Participation.objects.filter(match=match).delete()

        for i in range(1, 16):
            nom_d = (request.POST.get(f'nom_dom_{i}') or '').strip()
            prenom_d = (request.POST.get(f'prenom_dom_{i}') or '').strip()
            if nom_d or prenom_d:
                Participation.objects.create(
                    match=match,
                    numero_bonnet=i,
                    equipe_concernee='DOM',
                    nom=nom_d.upper(),
                    prenom=prenom_d.capitalize(),
                    numero_licence=request.POST.get(f'licence_dom_{i}', ''),
                    annee_naissance=_parse_int(request.POST.get(f'naissance_dom_{i}')),
                )

            nom_e = (request.POST.get(f'nom_ext_{i}') or '').strip()
            prenom_e = (request.POST.get(f'prenom_ext_{i}') or '').strip()
            if nom_e or prenom_e:
                Participation.objects.create(
                    match=match,
                    numero_bonnet=i,
                    equipe_concernee='EXT',
                    nom=nom_e.upper(),
                    prenom=prenom_e.capitalize(),
                    numero_licence=request.POST.get(f'licence_ext_{i}', ''),
                    annee_naissance=_parse_int(request.POST.get(f'naissance_ext_{i}')),
                )

        # Sauvegarde staff
        match.entraineur_dom = request.POST.get('entraineur_dom', '')
        match.entraineur_adj_dom = request.POST.get('entraineur_adj_dom', '')
        match.suppleant_dom = request.POST.get('suppleant_dom', '')
        match.entraineur_ext = request.POST.get('entraineur_ext', '')
        match.entraineur_adj_ext = request.POST.get('entraineur_adj_ext', '')
        match.suppleant_ext = request.POST.get('suppleant_ext', '')

        # Mise à jour catalogue si demandé
        if request.POST.get('update_master_dom') == 'on' and match.equipe_domicile_origine:
            match.equipe_domicile_origine.effectif.all().delete()
            for i in range(1, 16):
                nom = (request.POST.get(f'nom_dom_{i}') or '').strip()
                prenom = (request.POST.get(f'prenom_dom_{i}') or '').strip()
                if nom:
                    Joueur.objects.create(
                        equipe=match.equipe_domicile_origine,
                        nom=nom, prenom=prenom, numero_habituel=i,
                        numero_licence=request.POST.get(f'licence_dom_{i}', ''),
                        annee_naissance=_parse_int(request.POST.get(f'naissance_dom_{i}')),
                    )

        if request.POST.get('update_master_ext') == 'on' and match.equipe_exterieur_origine:
            match.equipe_exterieur_origine.effectif.all().delete()
            for i in range(1, 16):
                nom = (request.POST.get(f'nom_ext_{i}') or '').strip()
                prenom = (request.POST.get(f'prenom_ext_{i}') or '').strip()
                if nom:
                    Joueur.objects.create(
                        equipe=match.equipe_exterieur_origine,
                        nom=nom, prenom=prenom, numero_habituel=i,
                        numero_licence=request.POST.get(f'licence_ext_{i}', ''),
                        annee_naissance=_parse_int(request.POST.get(f'naissance_ext_{i}')),
                    )

        match.compo_validee = True
        match.save()
        return redirect('dashboard_arbitre', match_id=match.id)

    return render(request, 'gestion/compo_equipe.html', {
        'match': match,
        'slots_dom': _build_slots(match.equipe_domicile_origine),
        'slots_ext': _build_slots(match.equipe_exterieur_origine),
    })


def _parse_int(value):
    """Retourne un entier ou None si la valeur est absente/invalide."""
    try:
        return int(value) if value and str(value).strip() else None
    except (ValueError, TypeError):
        return None


# ==========================================
# 3. DASHBOARD ARBITRE
# ==========================================

def dashboard_arbitre(request, match_id):
    match = get_object_or_404(Match, id=match_id)
    joueurs_dom = Participation.objects.filter(
        match=match, equipe_concernee='DOM').order_by('numero_bonnet')
    joueurs_ext = Participation.objects.filter(
        match=match, equipe_concernee='EXT').order_by('numero_bonnet')
    events = Evenement.objects.filter(match=match).order_by('-heure_creation')
    tm_qs = Evenement.objects.filter(match=match, type_action='TM')
    tm_dom = tm_qs.filter(equipe_attribuee='DOM').count()
    tm_ext = tm_qs.filter(equipe_attribuee='EXT').count()
    return render(request, 'gestion/dashboard_arbitre.html', {
        'match': match,
        'joueurs_dom': joueurs_dom,
        'joueurs_ext': joueurs_ext,
        'events': events,
        'range_fautes': range(1, match.max_fautes_perso + 1),
        'tm_dom': tm_dom,
        'tm_ext': tm_ext,
    })


# ==========================================
# 4. API MATCH (cerveau du chrono & actions)
# ==========================================

def _chrono_payload(match):
    """Bloc chrono renvoyé à chaque réponse API (source de vérité)."""
    now_ts = timezone.now().timestamp()
    started_at = (match.dernier_top_chrono.timestamp()
                  if match.dernier_top_chrono else now_ts)
    return {
        'running':     match.chrono_en_cours,
        'frozen_time': float(match.temps_restant),
        'started_at':  started_at,
        'server_ts':   now_ts,
    }


def _shot_payload(match):
    """Bloc shot clock renvoyé à chaque réponse API (source de vérité).
    Si shot_top est None, on renvoie running=False — pas de started_at=now
    qui ferait croire au client que le chrono vient de démarrer (boucle)."""
    now_ts = timezone.now().timestamp()
    running = match.shot_en_cours and match.shot_top is not None
    started_at = match.shot_top.timestamp() if match.shot_top else 0.0
    return {
        'running':     running,
        'frozen_time': float(match.shot_restant),
        'started_at':  started_at,
        'server_ts':   now_ts,
        'max':         match.temps_possession,
    }


def _reset_shot(match, value=None):
    """Remet le shot clock (modifie l'objet, pas de save).
    value=None → temps_possession (28s standard)
    value=18   → 18s (coin, rebond offensif)"""
    match.shot_restant = value if value is not None else match.temps_possession
    if match.chrono_en_cours:
        match.shot_en_cours = True
        match.shot_top = timezone.now()
    else:
        match.shot_en_cours = False
        match.shot_top = None


def _stop_clocks(match):
    """Arrête match clock ET shot clock (sauvegardes valeurs courantes).
    Modifie l'objet, pas de save — appelé avant chaque action de jeu."""
    now = timezone.now()
    if match.chrono_en_cours:
        if match.dernier_top_chrono:
            delta = (now - match.dernier_top_chrono).total_seconds()
            match.temps_restant = max(0, int(match.temps_restant - delta))
        match.chrono_en_cours = False
        match.dernier_top_chrono = None
    if match.shot_en_cours and match.shot_top:
        elapsed = (now - match.shot_top).total_seconds()
        match.shot_restant = max(0, int(match.shot_restant - elapsed))
    match.shot_en_cours = False
    match.shot_top = None


def _history_qs(match):
    return list(
        Evenement.objects
        .filter(match=match)
        .order_by('-heure_creation')[:20]
        .values(
            'id', 'type_action', 'chrono_match', 'equipe_attribuee', 'periode',
            'score_dom_apres', 'score_ext_apres',
            'joueur__nom', 'joueur__numero_bonnet',
        )
    )


@csrf_exempt
def api_match_action(request, match_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    match = get_object_or_404(Match, id=match_id)
    action = data.get('action')

    # ── Chrono ───────────────────────────────────────────────────────────────

    if action == 'start_timer':
        if not match.chrono_en_cours:
            now = timezone.now()
            match.chrono_en_cours = True
            match.dernier_top_chrono = now
            if not match.shot_en_cours or match.shot_top is None:
                match.shot_en_cours = True
                match.shot_top = now
            match.save(update_fields=['chrono_en_cours', 'dernier_top_chrono',
                                       'shot_en_cours', 'shot_top'])
        return JsonResponse({'status': 'ok', 'chrono': _chrono_payload(match),
                             'shot': _shot_payload(match)})

    elif action == 'stop_timer':
        if match.chrono_en_cours:
            now = timezone.now()
            if match.dernier_top_chrono:
                delta = (now - match.dernier_top_chrono).total_seconds()
                match.temps_restant = max(0, int(match.temps_restant - delta))
            match.chrono_en_cours = False
            match.dernier_top_chrono = None
            if match.shot_en_cours and match.shot_top:
                elapsed = (now - match.shot_top).total_seconds()
                match.shot_restant = max(0, int(match.shot_restant - elapsed))
            match.shot_en_cours = False
            match.shot_top = None
            match.save(update_fields=['chrono_en_cours', 'temps_restant', 'dernier_top_chrono',
                                       'shot_en_cours', 'shot_restant', 'shot_top'])
        return JsonResponse({'status': 'ok', 'chrono': _chrono_payload(match),
                             'shot': _shot_payload(match)})

    elif action == 'adjust_time':
        try:
            delta = int(data.get('delta', 0))
        except (ValueError, TypeError):
            delta = 0
        delta = max(-600, min(600, delta))  # borne ±10 minutes max
        match.temps_restant = max(0, match.temps_restant + delta)
        match.save(update_fields=['temps_restant'])
        return JsonResponse({'status': 'ok', 'chrono': _chrono_payload(match),
                             'shot': _shot_payload(match)})

    elif action == 'shot_reset':
        # value=28 → possession neutre ; value=18 → situation offensive
        try:
            val = int(data.get('value', 0)) or None
        except (ValueError, TypeError):
            val = None
        _reset_shot(match, val)
        match.save(update_fields=['shot_en_cours', 'shot_restant', 'shot_top'])
        return JsonResponse({'status': 'ok', 'chrono': _chrono_payload(match),
                             'shot': _shot_payload(match)})

    elif action == 'shot_expired':
        # Shot clock = 0 : klaxon déjà joué côté client.
        # → arrêt du match clock + remise shot à 28s (prise de possession adverse)
        now = timezone.now()
        if match.chrono_en_cours:
            if match.dernier_top_chrono:
                delta = (now - match.dernier_top_chrono).total_seconds()
                match.temps_restant = max(0, int(match.temps_restant - delta))
            match.chrono_en_cours = False
            match.dernier_top_chrono = None
        match.shot_restant = match.temps_possession
        match.shot_en_cours = False
        match.shot_top = None
        match.save(update_fields=['chrono_en_cours', 'temps_restant', 'dernier_top_chrono',
                                   'shot_en_cours', 'shot_restant', 'shot_top'])
        return JsonResponse({'status': 'ok', 'chrono': _chrono_payload(match),
                             'shot': _shot_payload(match)})

    elif action == 'reset_period':
        match.chrono_en_cours = False
        match.dernier_top_chrono = None
        match.temps_restant = match.duree_periode * 60
        _reset_shot(match)
        match.save(update_fields=['chrono_en_cours', 'dernier_top_chrono', 'temps_restant',
                                   'shot_en_cours', 'shot_restant', 'shot_top'])
        return JsonResponse({'status': 'ok', 'chrono': _chrono_payload(match),
                             'shot': _shot_payload(match)})

    elif action == 'next_period':
        ScorePeriode.objects.update_or_create(
            match=match,
            numero_periode=match.periode_actuelle,
            defaults={'score_dom': match.score_domicile,
                      'score_ext': match.score_exterieur},
        )
        match.chrono_en_cours = False
        match.dernier_top_chrono = None
        match.periode_actuelle += 1
        match.temps_restant = match.duree_periode * 60
        _reset_shot(match)
        match.save(update_fields=[
            'chrono_en_cours', 'dernier_top_chrono',
            'periode_actuelle', 'temps_restant',
            'shot_en_cours', 'shot_restant', 'shot_top',
        ])
        return JsonResponse({
            'status': 'ok',
            'period': match.periode_actuelle,
            'chrono': _chrono_payload(match),
            'shot':   _shot_payload(match),
        })

    # ── Retour en jeu ─────────────────────────────────────────────────────────

    elif action == 'return_play':
        p_id = data.get('participation_id')
        joueur = get_object_or_404(Participation, id=p_id, match=match)
        joueur.est_exclu = False
        joueur.fin_exclusion = None
        joueur.save(update_fields=['est_exclu', 'fin_exclusion'])
        return JsonResponse({'status': 'ok', 'chrono': _chrono_payload(match)})

    # ── Présence terrain ──────────────────────────────────────────────────────

    elif action == 'toggle_sur_terrain':
        p_id = data.get('participation_id')
        joueur = get_object_or_404(Participation, id=p_id, match=match)
        joueur.est_sur_terrain = not joueur.est_sur_terrain
        joueur.save(update_fields=['est_sur_terrain'])
        return JsonResponse({'status': 'ok', 'est_sur_terrain': joueur.est_sur_terrain})

    # ── Actions de jeu ────────────────────────────────────────────────────────

    elif action == 'TM':
        chrono_display = data.get('chrono_display', '00:00')
        # equipe peut venir directement ou être dérivée du joueur (side panel)
        equipe = data.get('equipe')
        if not equipe and data.get('participation_id'):
            j = Participation.objects.filter(id=data['participation_id'], match=match).first()
            if j:
                equipe = j.equipe_concernee
        if not equipe:
            return JsonResponse({'status': 'error', 'message': 'Équipe non identifiée'})
        tm_count = Evenement.objects.filter(
            match=match, type_action='TM', equipe_attribuee=equipe).count()
        if tm_count >= match.nb_temps_morts:
            return JsonResponse({'status': 'error',
                                 'message': 'Temps morts épuisés pour cette équipe'})
        Evenement.objects.create(
            match=match, type_action='TM',
            chrono_match=chrono_display, periode=match.periode_actuelle,
            equipe_attribuee=equipe,
            score_dom_apres=match.score_domicile,
            score_ext_apres=match.score_exterieur,
        )

    elif action in ('B', 'E', 'EDA', 'EDAP', 'P', 'A', 'R', 'CJ€', 'CR'):
        p_id = data.get('participation_id')
        chrono_display = data.get('chrono_display', '00:00')

        joueur = None
        if p_id:
            joueur = get_object_or_404(Participation, id=p_id, match=match)

        # Champs chrono à sauvegarder quand les pendules s'arrêtent
        _CLOCK_FIELDS = [
            'chrono_en_cours', 'temps_restant', 'dernier_top_chrono',
            'shot_en_cours', 'shot_restant', 'shot_top',
        ]

        if action == 'B':
            # But → score + arrêt jeu + reset possession à 28s
            if joueur.equipe_concernee == 'DOM':
                match.score_domicile += 1
            else:
                match.score_exterieur += 1
            _stop_clocks(match)          # arrêt match clock + shot clock
            _reset_shot(match)           # shot à 28s (frozen car chrono=False)
            match.save(update_fields=['score_domicile', 'score_exterieur'] + _CLOCK_FIELDS)

        elif action == 'P':
            # Penalty → arrêt + faute personnelle
            joueur.nb_fautes_personnelles += 1
            if joueur.nb_fautes_personnelles >= match.max_fautes_perso:
                joueur.est_exclu_definitif = True
            joueur.save(update_fields=['nb_fautes_personnelles', 'est_exclu_definitif'])
            _stop_clocks(match)
            match.save(update_fields=_CLOCK_FIELDS)

        elif action == 'E':
            # Exclusion → arrêt + timer exclusion + shot clock gelé (reset manuel 18/28)
            joueur.nb_fautes_personnelles += 1
            joueur.est_exclu = True
            joueur.fin_exclusion = timezone.now() + timedelta(seconds=match.duree_exclusion)
            if joueur.nb_fautes_personnelles >= match.max_fautes_perso:
                joueur.est_exclu_definitif = True
                joueur.est_exclu = False
            joueur.save(update_fields=[
                'nb_fautes_personnelles', 'est_exclu',
                'fin_exclusion', 'est_exclu_definitif',
            ])
            _stop_clocks(match)
            match.save(update_fields=_CLOCK_FIELDS)

        elif action in ('EDA', 'EDAP'):
            # Exclusion définitive → arrêt + joueur hors jeu définitivement
            joueur.est_exclu_definitif = True
            joueur.est_exclu = False
            joueur.nb_fautes_personnelles += 1
            joueur.save(update_fields=[
                'est_exclu_definitif', 'est_exclu', 'nb_fautes_personnelles',
            ])
            _stop_clocks(match)
            match.save(update_fields=_CLOCK_FIELDS)

        elif action == 'CR':
            # Carton Rouge → exclusion définitive + arrêt
            if joueur:
                joueur.est_exclu_definitif = True
                joueur.est_exclu = False
                joueur.save(update_fields=['est_exclu_definitif', 'est_exclu'])
            _stop_clocks(match)
            match.save(update_fields=_CLOCK_FIELDS)

        elif action in ('A', 'R'):
            # Accident / Réclamation → arrêt neutre (pas de sanction)
            _stop_clocks(match)
            match.save(update_fields=_CLOCK_FIELDS)

        # CJ€ : avertissement équipe — pas d'arrêt ni exclusion (purement administratif)

        Evenement.objects.create(
            match=match, type_action=action,
            chrono_match=chrono_display, periode=match.periode_actuelle,
            joueur=joueur,
            equipe_attribuee=joueur.equipe_concernee if joueur else data.get('equipe'),
            score_dom_apres=match.score_domicile,
            score_ext_apres=match.score_exterieur,
        )

    # ── Annulation ───────────────────────────────────────────────────────────

    elif action == 'delete_event':
        evt_id = data.get('event_id')
        evt = get_object_or_404(Evenement, id=evt_id, match=match)

        if evt.type_action == 'B':
            if evt.equipe_attribuee == 'DOM':
                match.score_domicile = max(0, match.score_domicile - 1)
            else:
                match.score_exterieur = max(0, match.score_exterieur - 1)
            match.save(update_fields=['score_domicile', 'score_exterieur'])

        elif evt.type_action in ('E', 'EDA', 'EDAP', 'P') and evt.joueur:
            j = evt.joueur
            j.nb_fautes_personnelles = max(0, j.nb_fautes_personnelles - 1)
            if evt.type_action == 'E':
                j.est_exclu = False
                j.fin_exclusion = None
                if j.nb_fautes_personnelles < match.max_fautes_perso:
                    j.est_exclu_definitif = False
            elif evt.type_action in ('EDA', 'EDAP'):
                j.est_exclu_definitif = False
            j.save()

        evt.delete()

    # ── Refresh (état courant sans action) ────────────────────────────────────

    tm_qs = Evenement.objects.filter(match=match, type_action='TM')

    # Inclure l'état du joueur si une action vient de le modifier
    joueur_state = None
    _joueur = locals().get('joueur')
    if _joueur:
        joueur_state = {
            'id':              _joueur.id,
            'est_exclu_definitif': _joueur.est_exclu_definitif,
            'est_exclu':       _joueur.est_exclu,
            'fin_exclusion_ts': (_joueur.fin_exclusion.timestamp()
                                 if _joueur.fin_exclusion else 0),
        }

    resp = {
        'status':         'ok',
        'score_dom':      match.score_domicile,
        'score_ext':      match.score_exterieur,
        'period':         match.periode_actuelle,
        'history':        _history_qs(match),
        'chrono':         _chrono_payload(match),
        'shot':           _shot_payload(match),
        'tm_dom':         tm_qs.filter(equipe_attribuee='DOM').count(),
        'tm_ext':         tm_qs.filter(equipe_attribuee='EXT').count(),
        'nb_temps_morts': match.nb_temps_morts,
    }
    if joueur_state:
        resp['joueur_state'] = joueur_state
    return JsonResponse(resp)


# ==========================================
# 5. SCOREBOARD PUBLIC
# ==========================================

def score_board(request, match_id):
    match = get_object_or_404(Match, id=match_id)
    return render(request, 'gestion/score_board.html', {
        'match': match,
        'range_15': range(1, 16),
    })


@csrf_exempt
def upload_sponsor(request, match_id):
    """Upload one or more sponsor images for a match."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    match = get_object_or_404(Match, id=match_id)
    count = Sponsor.objects.filter(match=match).count()
    created = []
    for idx, f in enumerate(request.FILES.getlist('sponsors')):
        s = Sponsor.objects.create(match=match, image=f, ordre=count + idx)
        created.append({'id': s.id, 'url': request.build_absolute_uri(s.image.url)})
    return JsonResponse({'created': created})


@csrf_exempt
def delete_sponsor(request, sponsor_id):
    """Delete a sponsor image."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    s = get_object_or_404(Sponsor, id=sponsor_id)
    s.image.delete(save=False)
    s.delete()
    return JsonResponse({'status': 'deleted'})


@csrf_exempt
def api_match_state(request, match_id):
    """Endpoint de polling pour le scoreboard public."""
    match = get_object_or_404(Match, id=match_id)

    players_state = {}
    for p in Participation.objects.filter(match=match):
        if p.est_exclu_definitif:
            state, end_time = 'EDA', 0
        elif p.est_exclu and p.fin_exclusion:
            state = 'EXCL'
            end_time = p.fin_exclusion.timestamp()
        else:
            state, end_time = 'OK', 0
        players_state[p.id] = {
            'state':   state,
            'end_time': end_time,
            'fautes':  p.nb_fautes_personnelles,
            'side':    p.equipe_concernee,
            'bonnet':  p.numero_bonnet,
        }

    # Temps morts utilisés par équipe
    tm_qs = Evenement.objects.filter(match=match, type_action='TM')
    tm_dom = tm_qs.filter(equipe_attribuee='DOM').count()
    tm_ext = tm_qs.filter(equipe_attribuee='EXT').count()

    # Logos et sponsors
    request_host = request.build_absolute_uri('/')[:-1]
    logo_dom_url = (request_host + match.logo_dom.url) if match.logo_dom else None
    logo_ext_url = (request_host + match.logo_ext.url) if match.logo_ext else None
    sponsor_urls = [
        request_host + s.image.url
        for s in Sponsor.objects.filter(match=match)
    ]

    return JsonResponse({
        'score_dom':      match.score_domicile,
        'score_ext':      match.score_exterieur,
        'periode':        match.periode_actuelle,
        'players':        players_state,
        'chrono':         _chrono_payload(match),
        'tm_dom':         tm_dom,
        'tm_ext':         tm_ext,
        'nb_temps_morts': match.nb_temps_morts,
        'temps_possession': match.temps_possession,
        'nom_dom':        match.nom_equipe_domicile,
        'nom_ext':        match.nom_equipe_exterieur,
        'logo_dom':       logo_dom_url,
        'logo_ext':       logo_ext_url,
        'sponsors':       sponsor_urls,
        'couleur_dom':    match.couleur_bonnet_dom,
        'couleur_ext':    match.couleur_bonnet_ext,
        'shot':           _shot_payload(match),
    })


# ==========================================
# 6. EXPORT EXCEL (Feuille de match FFN)
# ==========================================

def export_excel(request, match_id):
    """
    Génère et télécharge la feuille de match officielle FFN en .xlsx.
    Mapping précis calé sur la structure réelle du template FFN.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return HttpResponse(
            "❌ Module manquant : openpyxl\n\npip install openpyxl",
            status=500,
            content_type='text/plain; charset=utf-8',
        )

    match = get_object_or_404(Match, id=match_id)

    # Accepte .xlsm (avec macros) en priorité, puis .xlsx en fallback
    base_dir = os.path.join(os.path.dirname(__file__), 'static', 'gestion')
    xlsm_path = os.path.join(base_dir, 'feuille_match_template.xlsm')
    xlsx_path  = os.path.join(base_dir, 'feuille_match_template.xlsx')
    if os.path.exists(xlsm_path):
        template_path = xlsm_path
        is_xlsm = True
    elif os.path.exists(xlsx_path):
        template_path = xlsx_path
        is_xlsm = False
    else:
        return HttpResponse(
            f'Template Excel introuvable.\n'
            f'Chemin attendu : {xlsm_path} (ou .xlsx)\n'
            'Copiez le template dans gestion/static/gestion/ '
            'sous le nom feuille_match_template.xlsm',
            status=500, content_type='text/plain; charset=utf-8',
        )

    wb = load_workbook(template_path, keep_vba=is_xlsm)
    ws = wb['Feuile de Match']   # typo intentionnelle du template FFN

    # Alias local pour éviter de répéter ws à chaque appel
    def w(row, col, value):
        _safe_write(ws, row, col, value)

    # ── Infos générales ──────────────────────────────────────────────────────
    # Nom équipe 1 : C3 (fusion C3:Q3)
    w(3, 3,  match.nom_equipe_domicile)
    # Compétition : AK3 (fusion AK3:AP3)
    w(3, 37, match.competition)
    # Lieu : AI4
    w(4, 35, match.lieu)
    # Date : AI5 (écrase la formule =TODAY() du template)
    w(5, 35, match.date_match.strftime('%d/%m/%Y'))
    # Heure : AN5
    if match.heure_debut:
        w(5, 40, match.heure_debut.strftime('%H:%M'))
    # Nom équipe 2 : C29 (fusion C29:Q29)
    w(29, 3, match.nom_equipe_exterieur)

    # ── Officiels – haut droite (lignes 8-10) ────────────────────────────────
    # Ligne 8 : Secrétaire (AH8) + IUF (AN8) | Chrono (AQ8) + IUF (AW8)
    w(8, 34, f"SECRETAIRE : {match.secretaire_nom}" if match.secretaire_nom else "SECRETAIRE :")
    w(8, 40, f"IUF N° : {match.secretaire_iuf}" if match.secretaire_iuf else "IUF N° :")
    w(8, 43, f"CHRONO : {match.chrono_nom}" if match.chrono_nom else "CHRONO :")
    w(8, 49, f"IUF N° : {match.chrono_iuf}" if match.chrono_iuf else "IUF N° :")
    # Ligne 9 : règles 30'/20' (AQ9)
    w(9, 43, f"30' / 20' : {match.temps_possession}'' / {match.duree_exclusion}''")
    # Ligne 10 : Juge de but 1 (AH10) + IUF (AN10) | Juge de but 2 (AQ10) + IUF (AW10)
    w(10, 34, f"JUGE DE BUT : {match.juge_but1_nom}" if match.juge_but1_nom else "JUGE DE BUT :")
    w(10, 40, f"IUF N° : {match.juge_but1_iuf}" if match.juge_but1_iuf else "IUF N° :")
    w(10, 43, f"JUGE DE BUT : {match.juge_but2_nom}" if match.juge_but2_nom else "JUGE DE BUT :")
    w(10, 49, f"IUF N° : {match.juge_but2_iuf}" if match.juge_but2_iuf else "IUF N° :")

    # ── Équipe domicile – joueurs (lignes 9-23, bonnet 1→15) ─────────────────
    joueurs_dom = list(
        Participation.objects.filter(match=match, equipe_concernee='DOM')
        .order_by('numero_bonnet')
    )
    _fill_players(ws, joueurs_dom, start_row=9, match=match)

    # ── Staff domicile (lignes 24-26, colonne C=3) ────────────────────────────
    w(24, 3, match.entraineur_dom)
    w(25, 3, match.entraineur_adj_dom)
    w(26, 3, match.suppleant_dom)

    # ── Temps morts : slots chronologiques, format "X {période}" ─────────────
    # DOM : ligne 5, colonnes I(9) J(10) K(11) L(12)
    # EXT : ligne 31, mêmes colonnes
    tm_slots = [9, 10, 11, 12]
    tm_events_dom = list(
        Evenement.objects
        .filter(match=match, type_action='TM', equipe_attribuee='DOM')
        .order_by('heure_creation')
    )
    tm_events_ext = list(
        Evenement.objects
        .filter(match=match, type_action='TM', equipe_attribuee='EXT')
        .order_by('heure_creation')
    )
    for i, ev in enumerate(tm_events_dom[:len(tm_slots)]):
        w(5,  tm_slots[i], f"X {ev.periode}")
    for i, ev in enumerate(tm_events_ext[:len(tm_slots)]):
        w(31, tm_slots[i], f"X {ev.periode}")

    # ── Résultats par période : delta (pas cumulatif) ─────────────────────────
    scores = {sp.numero_periode: sp
              for sp in ScorePeriode.objects.filter(match=match)}
    score_rows = {1: 26, 2: 27, 3: 29, 4: 30}
    prev_dom, prev_ext = 0, 0
    for p_num in sorted(score_rows.keys()):
        sp = scores.get(p_num)
        p_row = score_rows[p_num]
        if sp:
            w(p_row, 30, sp.score_dom - prev_dom)
            w(p_row, 31, sp.score_ext - prev_ext)
            prev_dom = sp.score_dom
            prev_ext = sp.score_ext

    # Dernier tronçon : si la période finale n'a pas de ScorePeriode (cas habituel
    # pour la P4 car on ne passe jamais à une P5), on écrit le delta restant.
    delta_dom = match.score_domicile - prev_dom
    delta_ext = match.score_exterieur - prev_ext
    if delta_dom or delta_ext:
        for p_num in sorted(score_rows.keys()):
            if p_num not in scores:
                p_row = score_rows[p_num]
                if delta_dom:
                    w(p_row, 30, delta_dom)
                if delta_ext:
                    w(p_row, 31, delta_ext)
                break

    # Score final – zone « RESULTAT FINAL » (R3C31 DOM, R5C31 EXT)
    w(3, 31, match.score_domicile)
    w(5, 31, match.score_exterieur)

    # ── Équipe extérieure – joueurs (lignes 35-49, bonnet 1→15) ──────────────
    joueurs_ext = list(
        Participation.objects.filter(match=match, equipe_concernee='EXT')
        .order_by('numero_bonnet')
    )
    _fill_players(ws, joueurs_ext, start_row=35, match=match)

    # ── Staff extérieur (lignes 50-52, colonne C=3) ───────────────────────────
    w(50, 3, match.entraineur_ext)
    w(51, 3, match.entraineur_adj_ext)
    w(52, 3, match.suppleant_ext)

    # ── Officiels – signatures (lignes 41-45) ─────────────────────────────────
    # Noms : col 37 (AK, fusion AK:AS)  |  IUF : col 46 (AT)
    w(41, 37, match.delegue_dom_nom)
    w(42, 37, match.delegue_ext_nom)
    w(43, 37, match.arbitre1_nom)
    w(43, 46, match.arbitre1_iuf)
    w(44, 37, match.arbitre2_nom)
    w(44, 46, match.arbitre2_iuf)
    w(45, 37, match.delegue_ffn_nom)
    w(45, 46, match.delegue_ffn_iuf)

    # ── Journal des événements (zone droite, 3 colonnes, lignes 13-39) ────────
    # 3 groupes démarrant aux colonnes 34 (AH), 40 (AN), 46 (AT)
    # Chaque groupe : TEMPS | B(dom) | N(ext) | CODE | SCORE
    # Max 27 lignes par groupe. 2 lignes grises séparent chaque période.
    from openpyxl.styles import PatternFill
    _gray = PatternFill('solid', fgColor='BBBBBB')

    events = list(
        Evenement.objects.filter(match=match).order_by('heure_creation')
    )
    groups = [(34, 13), (40, 13), (46, 13)]
    max_per_group = 27

    # Regrouper par période
    events_by_period = {}
    for evt in events:
        events_by_period.setdefault(evt.periode, []).append(evt)

    row_cursor = 0
    grp_idx = 0
    first_period = True

    for periode in sorted(events_by_period.keys()):
        # 2 lignes séparatrices grises entre chaque période
        if not first_period:
            for _ in range(2):
                if grp_idx >= len(groups):
                    break
                if row_cursor >= max_per_group:
                    grp_idx += 1
                    row_cursor = 0
                if grp_idx >= len(groups):
                    break
                sc, sr = groups[grp_idx]
                sep_row = sr + row_cursor
                for c in range(sc, sc + 5):
                    ws.cell(row=sep_row, column=c).fill = _gray
                row_cursor += 1
        first_period = False

        for evt in events_by_period[periode]:
            if grp_idx >= len(groups):
                break
            if row_cursor >= max_per_group:
                grp_idx += 1
                row_cursor = 0
            if grp_idx >= len(groups):
                break
            sc, sr = groups[grp_idx]
            row = sr + row_cursor

            bonnet = evt.joueur.numero_bonnet if evt.joueur else ''
            dom_b  = bonnet if evt.equipe_attribuee == 'DOM' else ''
            ext_b  = bonnet if evt.equipe_attribuee == 'EXT' else ''

            w(row, sc,     evt.chrono_match)
            w(row, sc + 1, dom_b)
            w(row, sc + 2, ext_b)
            w(row, sc + 3, evt.type_action)
            w(row, sc + 4, f"{evt.score_dom_apres}-{evt.score_ext_apres}")
            row_cursor += 1

    # ── Sauvegarde en mémoire et envoi HTTP ────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    equipes_slug = (
        f"{match.nom_equipe_domicile}_vs_{match.nom_equipe_exterieur}"
        .replace(' ', '_')[:60]
    )
    ext      = 'xlsm' if is_xlsm else 'xlsx'
    filename = f"feuille_match_{equipes_slug}.{ext}"
    mime     = (
        'application/vnd.ms-excel.sheet.macroEnabled.12'
        if is_xlsm else
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response = HttpResponse(buf.read(), content_type=mime)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _safe_write(ws, row, col, value):
    """
    Écrit dans une cellule en gérant les cellules fusionnées (MergedCell).
    Utilisable depuis _fill_players (qui n'a pas accès à la closure w()).
    """
    from openpyxl.cell.cell import MergedCell
    if value is None or value == '':
        return
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if (merged_range.min_row <= row <= merged_range.max_row
                    and merged_range.min_col <= col <= merged_range.max_col):
                cell = ws.cell(
                    row=merged_range.min_row,
                    column=merged_range.min_col,
                )
                break
        else:
            return
    cell.value = value


def _fill_players(ws, joueurs, start_row, match):
    """
    Remplit les lignes de joueurs dans le template FFN.
    start_row : 9 pour DOM, 35 pour EXT.

    Colonnes :
      B(2)=Licence  C(3)=Nom Prénom  P(16)=Naissance  Q(17)=Présence terrain
      R(18)=N°Bonnet
      S(19)=Buts P1  T(20)=Buts P2  U(21)=Buts P3  W(23)=Buts P4
      Z(26)+AA(27) = Code1+Période1
      AB(28)+AC(29) = Code2+Période2
      AD(30)+AE(31) = Code3+Période3
    """
    from django.db.models import Count

    joueur_ids = [p.id for p in joueurs]

    # Buts par joueur par période (2 requêtes au lieu de N)
    buts_per_period = {}
    for item in (
        Evenement.objects
        .filter(match=match, type_action='B', joueur_id__in=joueur_ids)
        .order_by()          # efface l'ordering par défaut (heure_creation) qui polluerait le GROUP BY
        .values('joueur_id', 'periode')
        .annotate(cnt=Count('id'))
    ):
        buts_per_period.setdefault(item['joueur_id'], {})[item['periode']] = item['cnt']

    # Sanctions par joueur (E, EDA, EDAP, P)
    sanctions = list(
        Evenement.objects
        .filter(match=match, type_action__in=('E', 'EDA', 'EDAP', 'P'), joueur_id__in=joueur_ids)
        .order_by('joueur_id', 'heure_creation')
    )
    events_map = {}
    for evt in sanctions:
        pid = evt.joueur_id
        if pid not in events_map:
            events_map[pid] = []
        if len(events_map[pid]) < 3:
            events_map[pid].append(evt)

    by_bonnet = {p.numero_bonnet: p for p in joueurs}

    for bonnet in range(1, 16):
        row = start_row + (bonnet - 1)
        p = by_bonnet.get(bonnet)
        if not p:
            continue

        _safe_write(ws, row, 2,  p.numero_licence or '')
        _safe_write(ws, row, 3,  f"{p.nom} {p.prenom}")
        _safe_write(ws, row, 16, p.annee_naissance or '')
        # Col Q(17) : présence sur le terrain (X = présent)
        if p.est_sur_terrain:
            _safe_write(ws, row, 17, 'X')
        _safe_write(ws, row, 18, bonnet)

        # Buts par période : S=P1(19), T=P2(20), U=P3(21), W=P4(23)
        period_cols = {1: 19, 2: 20, 3: 21, 4: 23}
        player_buts = buts_per_period.get(p.id, {})
        for period, col in period_cols.items():
            goals = player_buts.get(period, 0)
            if goals:
                _safe_write(ws, row, col, goals)

        # 3 slots de sanctions : (Code, Période) aux colonnes Z/AA, AB/AC, AD/AE
        code_cols = [(26, 27), (28, 29), (30, 31)]
        for evt_idx, (col_code, col_per) in enumerate(code_cols):
            player_evts = events_map.get(p.id, [])
            if evt_idx < len(player_evts):
                evt = player_evts[evt_idx]
                _safe_write(ws, row, col_code, evt.type_action)
                _safe_write(ws, row, col_per,  evt.periode)


# ==========================================
# 7. TABLEAU DE BORD (Legacy / fallback)
# ==========================================

def tableau_bord(request, match_id):
    match = get_object_or_404(Match, id=match_id)
    joueurs_dom = Participation.objects.filter(
        match=match, equipe_concernee='DOM').order_by('numero_bonnet')
    joueurs_ext = Participation.objects.filter(
        match=match, equipe_concernee='EXT').order_by('numero_bonnet')
    events = Evenement.objects.filter(match=match).order_by('-heure_creation')[:5]
    return render(request, 'gestion/tableau_bord.html', {
        'match': match,
        'joueurs_dom': joueurs_dom,
        'joueurs_ext': joueurs_ext,
        'events': events,
    })